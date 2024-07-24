import pandas as pd
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import time
from collections import defaultdict
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

URL_COLUMNS = {
    "Logo": "Logo URL",
    "Coaches": "2024 Coaches URL",
    "Majors": "Undergraduate Majors URL",
    "Roster": "2024 Roster URL",
    "Staff Dir": "Staff Directory"
}

def requests_retry_session(
    retries=3,
    backoff_factor=1,
    status_forcelist=(500, 502, 504),
    session=None,
):
    session = session or requests.Session()
    retry = Retry(
        total=retries,
        read=retries,
        connect=retries,
        backoff_factor=backoff_factor,
        status_forcelist=status_forcelist,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount('http://', adapter)
    session.mount('https://', adapter)
    return session

def check_url(url, school, category):
    if pd.isna(url) or url.strip() == '':
        return f"{school} - {category}: No URL provided"
    
    session = requests_retry_session()
    
    for attempt in range(3):  # 3 attempts total
        timeout = 2 * (2 ** attempt)  # 3s, then 6s, then 12s
        try:
            response = session.get(url, timeout=timeout)
            if response.status_code == 404:
                return f"{school} - {category}: 404 Error - {url}"
            elif response.status_code != 200:
                return f"{school} - {category}: HTTP {response.status_code} - {url}"
            return None  # Success
        except requests.Timeout:
            if attempt == 2:  # Last attempt
                return f"{school} - {category}: Timeout (after {timeout}s) - {url}"
        except requests.RequestException as e:
            return f"{school} - {category}: Error - {str(e)} - {url}"
    
    return f"{school} - {category}: Unexpected error - {url}"

def process_sheet(sheet_name, df):
    errors = defaultdict(list)
    timeouts = defaultdict(list)
    with ThreadPoolExecutor(max_workers=20) as executor:
        futures = []
        for category, column in URL_COLUMNS.items():
            if column in df.columns:
                futures.extend([executor.submit(check_url, row[column], row['School'], category) 
                                for _, row in df.iterrows()])
        for future in as_completed(futures):
            result = future.result()
            if result:
                category = result.split(" - ")[1].split(":")[0]
                if "Timeout" in result:
                    timeouts[category].append(result)
                else:
                    errors[category].append(result)
    return errors, timeouts

def highlight_broken_urls(wb, sheet_name, broken_urls):
    light_red_fill = PatternFill(start_color='FFFF9999', end_color='FFFF9999', fill_type='solid')
    ws = wb[sheet_name]
    for category, column_name in URL_COLUMNS.items():
        for cell in ws[1]:
            if cell.value == column_name:
                col_letter = cell.column_letter
                for row in ws[col_letter]:
                    if row.value in broken_urls:
                        row.fill = light_red_fill
    return wb

def clear_past_results(txt_file, xlsx_file):
    for file in [txt_file, xlsx_file]:
        if os.path.exists(file):
            os.remove(file)
            print(f"Removed previous {file}")

def write_errors_to_file(f, errors, timeouts, sheet_name):
    f.write(f"\n{sheet_name}\n")
    f.write("=" * len(sheet_name) + "\n")
    total_errors = 0
    total_timeouts = 0
    for category in URL_COLUMNS.keys():
        category_errors = sorted(errors[category], key=lambda x: x.split(" - ")[0])
        category_timeouts = sorted(timeouts[category], key=lambda x: x.split(" - ")[0])
        if category_errors or category_timeouts:
            f.write(f"\n{category}:\n")
            f.write("-" * len(category) + "\n")
            if category_errors:
                f.write("Errors:\n")
                for error in category_errors:
                    f.write(f"{error}\n")
                f.write(f"Total {category} errors: {len(category_errors)}\n")
                total_errors += len(category_errors)
            if category_timeouts:
                f.write("Timeouts:\n")
                for timeout in category_timeouts:
                    f.write(f"{timeout}\n")
                f.write(f"Total {category} timeouts: {len(category_timeouts)}\n")
                total_timeouts += len(category_timeouts)
    f.write(f"\nTotal errors in {sheet_name}: {total_errors}\n")
    f.write(f"Total timeouts in {sheet_name}: {total_timeouts}\n")
    return total_errors, total_timeouts

def main():
    input_file = r"C:\Users\dell3\source\repos\school-data-scraper-2\Freelancer_Data_Mining_Project.xlsx"
    output_file = "Freelancer_Data_Mining_Project_Highlighted.xlsx"
    error_file = "url_errors.txt"

    clear_past_results(error_file, output_file)

    xls = pd.ExcelFile(input_file)
    wb = load_workbook(input_file)
    
    all_errors = {}
    all_broken_urls = set()
    total_sheets = len(xls.sheet_names)
    
    start_time = time.time()

    with open(error_file, "w") as f:
        for index, sheet_name in enumerate(xls.sheet_names, 1):
            sheet_start_time = time.time()
            print(f"Processing sheet: {sheet_name} ({index}/{total_sheets})")
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            sheet_errors, sheet_timeouts = process_sheet(sheet_name, df)
            all_errors[sheet_name] = (sheet_errors, sheet_timeouts)
            
            total_errors, total_timeouts = write_errors_to_file(f, sheet_errors, sheet_timeouts, sheet_name)
            
            sheet_broken_urls = set()
            for category_errors in sheet_errors.values():
                for error in category_errors:
                    url = error.split(" - ")[-1]
                    all_broken_urls.add(url)
                    sheet_broken_urls.add(url)
            
            # Highlight broken URLs for this sheet
            wb = highlight_broken_urls(wb, sheet_name, sheet_broken_urls)
            
            # Save intermediate results
            wb.save(output_file)
            f.flush()  # Ensure the text file is updated
            
            sheet_end_time = time.time()
            sheet_duration = sheet_end_time - sheet_start_time
            print(f"Completed processing {sheet_name} in {sheet_duration:.2f} seconds")
            print(f"Intermediate results saved. Errors: {total_errors}, Timeouts: {total_timeouts}")
    
    end_time = time.time()
    total_duration = end_time - start_time
    
    print(f"\nAll Errors and Issues have been saved to {error_file}")
    print(f"Highlighted Excel file saved as {output_file}")
    
    print("\nOverall Statistics:")
    total_errors = sum(sum(len(errors) for errors in sheet_errors.values()) for sheet_errors, _ in all_errors.values())
    total_timeouts = sum(sum(len(timeouts) for timeouts in sheet_timeouts.values()) for _, sheet_timeouts in all_errors.values())
    print(f"Total broken URLs across all sheets: {total_errors}")
    print(f"Total timeouts across all sheets: {total_timeouts}")
    
    for sheet, (errors, timeouts) in all_errors.items():
        sheet_total_errors = sum(len(category_errors) for category_errors in errors.values())
        sheet_total_timeouts = sum(len(category_timeouts) for category_timeouts in timeouts.values())
        if sheet_total_errors > 0 or sheet_total_timeouts > 0:
            print(f"\n{sheet}:")
            if sheet_total_errors > 0:
                print(f"  Errors: {sheet_total_errors}")
                for category, category_errors in errors.items():
                    if category_errors:
                        print(f"    {category}: {len(category_errors)} errors")
            if sheet_total_timeouts > 0:
                print(f"  Timeouts: {sheet_total_timeouts}")
                for category, category_timeouts in timeouts.items():
                    if category_timeouts:
                        print(f"    {category}: {len(category_timeouts)} timeouts")
    
    print(f"\nTotal processing time: {total_duration:.2f} seconds")

if __name__ == "__main__":
    main()