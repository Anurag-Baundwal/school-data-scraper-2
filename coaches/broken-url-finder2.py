import pandas as pd
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

def check_url(url, school):
    if pd.isna(url):
        return f"{school}: No URL provided"
    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 404:
            return f"{school}: 404 Error - {url}"
    except requests.RequestException as e:
        return f"{school}: Error - {str(e)} - {url}"
    return None

def process_sheet(sheet_name, df):
    errors = []
    with ThreadPoolExecutor(max_workers=10) as executor:
        future_to_url = {executor.submit(check_url, row['2024 Coaches URL'], row['School']): row for _, row in df.iterrows()}
        for future in as_completed(future_to_url):
            result = future.result()
            if result:
                errors.append(result)
    return errors

def highlight_broken_urls(input_file, output_file, broken_urls):
    wb = load_workbook(input_file)
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        url_column = None
        
        # Find the '2024 Coaches URL' column
        for cell in ws[1]:
            if cell.value == '2024 Coaches URL':
                url_column = cell.column_letter
                break
        
        if url_column:
            for cell in ws[url_column]:
                if cell.value in broken_urls:
                    cell.fill = red_fill
    
    wb.save(output_file)

def clear_past_results(txt_file, xlsx_file):
    if os.path.exists(txt_file):
        os.remove(txt_file)
        print(f"Removed previous {txt_file}")
    
    if os.path.exists(xlsx_file):
        os.remove(xlsx_file)
        print(f"Removed previous {xlsx_file}")

def main():
    input_file = r"C:\Users\dell3\source\repos\school-data-scraper-2\Freelancer_Data_Mining_Project.xlsx"
    output_file = "Freelancer_Data_Mining_Project_Highlighted.xlsx"
    error_file = "url_errors.txt"

    clear_past_results(error_file, output_file)

    xls = pd.ExcelFile(input_file)
    
    all_errors = {}
    all_broken_urls = set()
    
    with open(error_file, "w") as f:
        for sheet_name in xls.sheet_names:
            print(f"Processing sheet: {sheet_name}")
            df = pd.read_excel(input_file, sheet_name=sheet_name)
            sheet_errors = process_sheet(sheet_name, df)
            all_errors[sheet_name] = sheet_errors
            
            f.write(f"\n{sheet_name}\n")
            f.write("=" * len(sheet_name) + "\n")
            for error in sheet_errors:
                f.write(f"{error}\n")
                all_broken_urls.add(error.split(" - ")[-1])
            
            f.write(f"\nTotal broken URLs in {sheet_name}: {len(sheet_errors)}\n")
            print(f"Completed processing {sheet_name}")
    
    print(f"\nAll 404 Errors and Issues have been saved to {error_file}")
    
    print("\nOverall Statistics:")
    total_broken = sum(len(errors) for errors in all_errors.values())
    print(f"Total broken URLs across all sheets: {total_broken}")
    for sheet, errors in all_errors.items():
        print(f"{sheet}: {len(errors)} broken URLs")
    
    print("\nCreating highlighted Excel file...")
    highlight_broken_urls(input_file, output_file, all_broken_urls)
    print(f"Highlighted Excel file saved as {output_file}")

if __name__ == "__main__":
    main()